import os
import sys
from datetime import datetime

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from bot import config, db, excel_export, logic  # noqa: E402


@pytest.fixture()
def conn(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "DB_PATH", str(tmp_path / "test.sqlite3"))
    monkeypatch.setattr(config, "EXCEL_LOCAL_DIR", str(tmp_path / "exports"))
    db.init_db()
    connection = db.get_connection()
    yield connection
    connection.close()


def dt(h, m, date="2026-08-14"):
    y, mo, d = (int(x) for x in date.split("-"))
    return datetime(y, mo, d, h, m)


def _seed(conn, date="2026-08-14"):
    logic.start_day(conn, date)
    logic.process_checkin(conn, "c", date, 1, dt(9, 0), "A 133\nS 160\nL3 80")
    logic.process_checkin(conn, "c", date, 1, dt(10, 0), "A 266\nS 320\nL3 160")


def test_export_creates_expected_sheets_and_banner(conn):
    _seed(conn)
    path = excel_export.export_workbook(conn, "2026-08-14")
    assert os.path.exists(path)

    wb = load_workbook(path)
    assert set(excel_export.GENERATED_SHEETS + ["Notes"]) <= set(wb.sheetnames)

    ws = wb["Donnees"]
    assert ws["A1"].value == excel_export.BANNER_TEXT
    assert [c.value for c in ws[2]] == excel_export.DONNEES_HEADERS
    assert ws.freeze_panes == "A3"

    data_rows = [row for row in ws.iter_rows(min_row=3, values_only=True) if row[0] is not None]
    assert len(data_rows) == 6  # 2 points x 3 lignes


def test_export_never_overwrites_notes_sheet(conn):
    _seed(conn)
    path = excel_export.export_workbook(conn, "2026-08-14")

    wb = load_workbook(path)
    wb["Notes"]["A5"] = "Commentaire humain a preserver"
    wb.save(path)

    logic.process_checkin(conn, "c", "2026-08-14", 1, dt(11, 0), "A 399")
    path = excel_export.export_workbook(conn, "2026-08-14")

    wb = load_workbook(path)
    assert wb["Notes"]["A5"].value == "Commentaire humain a preserver"


def test_export_regenerates_donnees_without_duplicating_rows(conn):
    _seed(conn)
    excel_export.export_workbook(conn, "2026-08-14")
    path = excel_export.export_workbook(conn, "2026-08-14")  # regenere sans nouvelle donnee

    wb = load_workbook(path)
    ws = wb["Donnees"]
    data_rows = [row for row in ws.iter_rows(min_row=3, values_only=True) if row[0] is not None]
    assert len(data_rows) == 6  # toujours 6, pas 12


def test_synthese_quotidienne_rows_reuse_compute_poste_stats(conn):
    _seed(conn)
    rows = excel_export.synthese_quotidienne_rows(conn, 2026)
    assert len(rows) == 3  # une ligne par code actif (A, S, SKD)
    codes_lignes = {r[2] for r in rows}
    assert codes_lignes == {"Ligne Auto", "Ligne Semi-auto", "Ligne 03"}

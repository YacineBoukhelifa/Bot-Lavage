"""Export Excel local (.xlsx par année) — openpyxl (spec v2 §7.1, §7.5a).

Le classeur ne régénère QUE les feuilles qu'il possède (`Donnees`,
`Synthese_Quotidienne`, `Vue_Mensuelle`) ; une feuille `Notes` est créée une
fois et jamais retouchée — c'est l'endroit prévu pour l'annotation humaine.
Ce module ne fait aucun appel réseau : Google Sheets/Drive (§7.4-7.5) sont
dans `bot.gsheets` / `bot.gdrive`.
"""
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import config, logic

GENERATED_SHEETS = ["Donnees", "Synthese_Quotidienne", "Vue_Mensuelle"]
BANNER_TEXT = "⚠️ Feuille générée automatiquement — toute modification manuelle sera écrasée. Écrire dans la feuille Notes."

JOURS_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]

DONNEES_HEADERS = [
    "Date production", "Jour", "Poste", "Ligne", "Code", "Point", "Créneau", "Coef",
    "Cumul", "Production", "Objectif", "Écart", "Écart %", "Saisi par", "Horodatage réel",
]
SYNTHESE_HEADERS = [
    "Date production", "Poste", "Ligne", "Total produit", "Objectif poste",
    "Écart", "Taux réalisation", "Moyenne horaire", "Nb points saisis", "Statut",
]


def workbook_path(year):
    os.makedirs(config.EXCEL_LOCAL_DIR, exist_ok=True)
    return os.path.join(config.EXCEL_LOCAL_DIR, f"Production_BU_Lavage_{year}.xlsx")


def _ensure_workbook(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Notes libres — cette feuille n'est jamais modifiée par le bot."
    notes["A1"].font = Font(italic=True)
    return wb


def _reset_sheet(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    return ws


def _write_banner_and_headers(ws, headers):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    banner = ws.cell(row=1, column=1, value=BANNER_TEXT)
    banner.font = Font(bold=True, color="9C5700")
    banner.fill = PatternFill("solid", fgColor="FFEB9C")
    banner.alignment = Alignment(horizontal="left")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    ws.freeze_panes = "A3"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _donnees_rows(conn, year):
    rows = conn.execute(
        "SELECT * FROM checkpoints WHERE date LIKE ? ORDER BY date, poste, code, heure",
        (f"{year}-%",),
    ).fetchall()
    out = []
    for r in rows:
        date_obj = datetime.strptime(r["date"], "%Y-%m-%d")
        creneau = f"{logic.heure_debut_periode(r['poste'], r['heure'])}-{r['heure']}"
        out.append([
            date_obj.strftime("%d/%m/%Y"),
            JOURS_FR[date_obj.weekday()],
            r["poste"],
            config.LIGNES[r["code"]]["nom_affiche"],
            r["code"],
            r["heure"],
            creneau,
            r["coef"],
            r["cumul_envoye"],
            r["production_heure"],
            r["objectif"],
            r["ecart"],
            r["ecart_pct"],
            r["saisi_par_nom"] or r["saisi_par_id"] or "-",
            r["horodatage_reel"],
        ])
    return out


def synthese_quotidienne_rows(conn, year):
    postes_du_marche = conn.execute(
        "SELECT DISTINCT date, poste FROM checkpoints WHERE date LIKE ? ORDER BY date, poste",
        (f"{year}-%",),
    ).fetchall()
    out = []
    for row in postes_du_marche:
        stats = logic.compute_poste_stats(conn, row["date"], row["poste"])
        date_obj = datetime.strptime(row["date"], "%Y-%m-%d")
        poste_row = logic.get_poste(conn, row["date"], row["poste"])
        statut = poste_row["statut"] if poste_row else "inconnu"
        for l in stats["lignes"]:
            nb_points = sum(1 for p in l["points"] if p["active"])
            out.append([
                date_obj.strftime("%d/%m/%Y"),
                row["poste"],
                l["nom"],
                l["total"],
                l["objectif"],
                l["ecart"],
                (round(l["total"] / l["objectif"] * 100, 1) if l["objectif"] else None),
                l["moyenne_horaire"],
                nb_points,
                statut,
            ])
    return out


def vue_mensuelle(conn, year, month):
    """Tableau croise du mois de `date` : une ligne par jour, une colonne par
    (point de controle x ligne) du poste 1, plus total/taux en fin de ligne."""
    prefix = f"{year}-{month:02d}-"
    dates = [
        r["date"] for r in conn.execute(
            "SELECT DISTINCT date FROM checkpoints WHERE date LIKE ? AND poste = 1 ORDER BY date",
            (f"{prefix}%",),
        ).fetchall()
    ]
    points_heures = [p["heure"] for p in config.POSTE_1["points"]]
    headers = ["Date"]
    for heure in points_heures:
        for code in config.ORDRE_AFFICHAGE:
            headers.append(f"{heure} {code}")
    headers += ["Total jour", "Taux réalisation"]

    rows = [headers]
    for date in dates:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        row = [date_obj.strftime("%d/%m/%Y")]
        for heure in points_heures:
            for code in config.ORDRE_AFFICHAGE:
                r = conn.execute(
                    "SELECT production_heure FROM checkpoints WHERE date=? AND poste=1 AND code=? AND heure=?",
                    (date, code, heure),
                ).fetchone()
                row.append(r["production_heure"] if r else None)
        stats = logic.compute_poste_stats(conn, date, 1)
        row.append(stats["total_general"])
        row.append(round(stats["taux_global_pct"], 1) if stats["taux_global_pct"] is not None else None)
        rows.append(row)
    return rows


def export_workbook(conn, date):
    """Régénère le classeur annuel (année de `date`) et retourne son chemin
    local. N'écrit jamais dans `Notes`."""
    year = int(date.split("-")[0])
    month = int(date.split("-")[1])
    path = workbook_path(year)
    wb = _ensure_workbook(path)

    ws = _reset_sheet(wb, "Donnees")
    _write_banner_and_headers(ws, DONNEES_HEADERS)
    for r_idx, row in enumerate(_donnees_rows(conn, year), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    ws = _reset_sheet(wb, "Synthese_Quotidienne")
    _write_banner_and_headers(ws, SYNTHESE_HEADERS)
    for r_idx, row in enumerate(synthese_quotidienne_rows(conn, year), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    mensuelle = vue_mensuelle(conn, year, month)
    ws = _reset_sheet(wb, "Vue_Mensuelle")
    _write_banner_and_headers(ws, mensuelle[0] if mensuelle else ["Date"])
    for r_idx, row in enumerate(mensuelle[1:], start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    if "Notes" not in wb.sheetnames:
        notes = wb.create_sheet("Notes")
        notes["A1"] = "Notes libres — cette feuille n'est jamais modifiée par le bot."
        notes["A1"].font = Font(italic=True)

    wb.save(path)
    return path
